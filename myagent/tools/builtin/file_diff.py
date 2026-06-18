"""file_diff builtin tool.

对比两个文件的差异：
- 文本类（txt/csv/docx/pdf）：按文本行做传统 diff（difflib）。
- XLSX vs XLSX：按"同名表"匹配后逐行对比；只在一侧存在的表返回概要。

设计要点：
- 参数极简，仅暴露 path_a / path_b / table_name / columns 给 LLM。
- 截断与告警阈值是模块常量，不进 schema，避免污染 LLM 可见参数。
- 大差异（> _WARN_THRESHOLD）在 content 顶部加告警，并按 _MAX_OUTPUT_CHARS 截断；
  LLM 通过 table_name / columns 收窄范围获取细节，而非翻页。
"""
import logging
from pathlib import Path
from typing import Any

from myagent.tools.api import ToolResult, tool
from myagent.tools.builtin._file_common import (
    _check_path_safety,
    _detect_encoding,
    _detect_file_type,
    _extract_docx_lines,
    _json_value,
    _parse_a1_range,
    _range_from_bounds,
    _worksheet_tables,
)

logger = logging.getLogger(__name__)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 输出控制常量（不暴露给 LLM）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_WARN_THRESHOLD = 1000        # 差异量超此值 → content 顶部加告警
_MAX_OUTPUT_CHARS = 12000     # content 截断上限（约 3~4k token）
_MAX_HUNKS_IN_CONTENT = 30    # 文本 diff 在 content 中展示的 hunk 上限
_MAX_CHANGES_PER_TABLE = 30   # 每个同名表在 content 中展示的单元格差异上限
_SAMPLE_ROWS = 8              # 无同名表概要的样本数据行数

_TEXT_TYPES = {"text", "csv", "docx", "pdf"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 主入口
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@tool(
    name="file_diff",
    timeout=120,
    description=(
        "对比两个文件的差异并返回受控的差异报告。\n"
        "- 文本类（txt/csv/docx/pdf）：按文本行做传统 diff（新增/删除/替换的行）。\n"
        "- XLSX vs XLSX：在两个文件中按名字匹配同名表（命中 sheet 名或原生 table 名均可），"
        "对匹配到的表逐行对比；只在一侧存在的表返回概要，不报错。\n"
        "大差异会自动截断并告警，可用 table_name/columns 收窄范围。"
    ),
)
async def file_diff(
    path_a: str,
    path_b: str,
    table_name: str | None = None,
    columns: list[int] = None,
) -> ToolResult:
    """
    对比两个文件差异。docx/pdf/txt/csv 按文本行 diff；xlsx 按"同名表"匹配后逐行对比。

    Args:
        path_a: 第一个文件路径。推荐传入绝对路径；相对路径请先用 workspace root 拼成绝对路径。本工具不自动解析相对路径
        path_b: 第二个文件路径。推荐传入绝对路径
        table_name: 可选（仅 xlsx 生效）。指定一个表名，在两个文件中按名字搜索该表并只对比它；默认 None=对比两文件中所有同名表。命中 sheet 名或原生 table 名均可
        columns: 可选（仅 xlsx 生效）。逐行对比时只比较这些列。传两个元素且 a<b 视为闭区间，如 [2,5] 表示第 2~5 列；传多个元素如 [2,4,6] 视为显式指定列。默认 None=比较全部列
    """
    # ── 参数规整 ──
    path_a = str(path_a).strip() if path_a else ""
    path_b = str(path_b).strip() if path_b else ""
    table_name = str(table_name).strip() or None if table_name else None
    columns = _normalize_columns(columns)

    if not path_a or not path_b:
        return ToolResult(content="path_a 和 path_b 都不能为空。", is_error=True)

    for label, p in (("path_a", path_a), ("path_b", path_b)):
        err = _check_path_safety(p)
        if err:
            return ToolResult(content=f"{label}: {err}", is_error=True)

    target_a, target_b = Path(path_a), Path(path_b)
    for label, t in (("path_a", target_a), ("path_b", target_b)):
        if not t.exists():
            return ToolResult(content=f"{label}: 文件不存在: {t}", is_error=True)
        if not t.is_file():
            return ToolResult(content=f"{label}: 不是文件: {t}", is_error=True)

    type_a = _detect_file_type(target_a)
    type_b = _detect_file_type(target_b)
    logger.info("file_diff 开始: a=%s(%s) b=%s(%s) table_name=%s columns=%s",
                target_a.name, type_a, target_b.name, type_b, table_name, columns)

    try:
        if type_a == "xlsx" and type_b == "xlsx":
            return await _diff_xlsx(target_a, target_b, table_name, columns)
        if type_a in _TEXT_TYPES and type_b in _TEXT_TYPES:
            return await _diff_text(target_a, target_b, type_a, type_b)

        return ToolResult(
            content=(
                f"两文件类型组合不支持: A={type_a}, B={type_b}。\n"
                "支持: 两个文本类(txt/csv/docx/pdf) 互相 diff；或 两个 xlsx 互相 diff。"
            ),
            is_error=True,
            metadata={"type_a": type_a, "type_b": type_b},
        )
    except Exception as exc:
        logger.exception("file_diff 执行异常")
        return ToolResult(
            content=f"file_diff 执行异常: {type(exc).__name__}: {exc}",
            is_error=True,
            metadata={"type": type(exc).__name__},
        )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 文本类 diff
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


async def _diff_text(
    target_a: Path, target_b: Path, type_a: str, type_b: str
) -> ToolResult:
    lines_a = _load_text_lines(target_a, type_a)
    lines_b = _load_text_lines(target_b, type_b)

    hunks, stats = _diff_text_lines(lines_a, lines_b)
    total = len(hunks)

    content = _render_text_report(target_a, target_b, type_a, type_b,
                                  len(lines_a), len(lines_b), stats, hunks)
    content, truncated = _truncate(content)

    metadata = {
        "mode": "text",
        "path_a": str(target_a.resolve()),
        "path_b": str(target_b.resolve()),
        "type_a": type_a,
        "type_b": type_b,
        "lines_a": len(lines_a),
        "lines_b": len(lines_b),
        "stats": stats,
        "hunk_count": total,
        "hunks": _compact_hunks(hunks),
        "truncated": truncated,
        "warn_threshold": _WARN_THRESHOLD,
    }
    content = _maybe_warn(content, total)
    logger.info("file_diff 文本完成: hunks=%d added=%d removed=%d",
                total, stats["added"], stats["removed"])
    return ToolResult(content=content, metadata=metadata)


def _load_text_lines(path: Path, file_type: str) -> list[str]:
    """把文本类文件解析成纯文本行列表（供 difflib 使用）。"""
    if file_type == "text":
        return _read_plain_lines(path)
    if file_type == "csv":
        return _read_csv_lines(path)
    if file_type == "docx":
        return _read_docx_lines(path)
    if file_type == "pdf":
        return _read_pdf_lines(path)
    raise ValueError(f"不支持的文本类类型: {file_type}")


def _read_plain_lines(path: Path) -> list[str]:
    enc = _detect_encoding(path)
    if enc == "binary":
        raise ValueError(f"{path.name} 似乎是二进制文件")
    with open(path, "r", encoding=enc, errors="replace") as f:
        return [line.rstrip("\n\r") for line in f.readlines()]


def _read_csv_lines(path: Path) -> list[str]:
    import csv as _csv

    enc = _detect_encoding(path)
    if enc == "binary":
        enc = "utf-8"
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, "r", encoding=enc, errors="replace", newline="") as f:
        rows = list(_csv.reader(f, delimiter=delimiter))
    return [" | ".join(str(cell) for cell in row) for row in rows]


def _read_docx_lines(path: Path) -> list[str]:
    if path.suffix.lower() == ".doc":
        raise ValueError("不支持旧版 .doc 格式，请先转换为 .docx")
    from docx import Document

    doc = Document(str(path))
    return [line.rstrip("\n\r") for line in _extract_docx_lines(doc)]


def _read_pdf_lines(path: Path) -> list[str]:
    # 抑制 pdfminer DEBUG 日志洪水（同 file_read/file_query 的处理）
    for name in ("pdfminer", "pdfminer.psparser", "pdfminer.pdfinterp",
                 "pdfminer.pdfpage", "pdfminer.converter",
                 "pdfminer.layout", "pdfminer.utils"):
        logging.getLogger(name).setLevel(logging.WARNING)

    import pdfplumber

    lines: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                lines.append(line.rstrip("\n\r"))
    return lines


def _diff_text_lines(lines_a: list[str], lines_b: list[str]) -> tuple[list[dict], dict]:
    """用 difflib 计算 hunk。返回 (hunks, stats)。行号 1-based。"""
    import difflib

    matcher = difflib.SequenceMatcher(a=lines_a, b=lines_b, autojunk=False)
    hunks: list[dict] = []
    added = removed = replaced = 0

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        hunks.append({
            "tag": tag,
            "a_start": i1 + 1, "a_end": i2,
            "b_start": j1 + 1, "b_end": j2,
            "a": lines_a[i1:i2],
            "b": lines_b[j1:j2],
        })
        if tag == "delete":
            removed += i2 - i1
        elif tag == "insert":
            added += j2 - j1
        elif tag == "replace":
            replaced += max(i2 - i1, j2 - j1)

    return hunks, {"added": added, "removed": removed, "replaced": replaced}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# XLSX diff
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


async def _diff_xlsx(
    target_a: Path, target_b: Path,
    table_name: str | None, columns: list[int] | None,
) -> ToolResult:
    try:
        import openpyxl
    except ImportError:
        return ToolResult(
            content="缺少 openpyxl 库。请安装: pip install openpyxl",
            is_error=True,
        )

    wb_a = openpyxl.load_workbook(str(target_a), read_only=True, data_only=True)
    wb_b = openpyxl.load_workbook(str(target_b), read_only=True, data_only=True)

    try:
        units_a = _collect_named_units(wb_a)
        units_b = _collect_named_units(wb_b)

        if table_name:
            focus = table_name
            in_a = focus in units_a
            in_b = focus in units_b
            if not (in_a or in_b):
                return ToolResult(
                    content=f"表名 {focus!r} 在两个文件中均未找到。\n"
                            f"A 的表: {sorted(units_a)}\nB 的表: {sorted(units_b)}",
                    is_error=True,
                    metadata={"mode": "xlsx", "table_name": focus,
                              "names_a": sorted(units_a), "names_b": sorted(units_b)},
                )
            common = {focus} if (in_a and in_b) else set()
            only_a = {focus} if (in_a and not in_b) else set()
            only_b = {focus} if (in_b and not in_a) else set()
        else:
            common = set(units_a) & set(units_b)
            only_a = set(units_a) - common
            only_b = set(units_b) - common

        matched: list[dict] = []
        for name in sorted(common):
            unit_a = _pick_unit(units_a[name])
            unit_b = _pick_unit(units_b[name])
            headers_a, rows_a = _unit_matrix(unit_a)
            headers_b, rows_b = _unit_matrix(unit_b)
            ncol = max(len(headers_a), len(headers_b), 1)
            cols = _resolve_columns(columns, ncol)
            changes = _diff_matrices(rows_a, rows_b, cols)
            matched.append({
                "name": name,
                "kind_a": unit_a["kind"], "kind_b": unit_b["kind"],
                "sheet_a": unit_a["sheet"], "sheet_b": unit_b["sheet"],
                "ref_a": _range_from_bounds(*unit_a["bounds"]),
                "ref_b": _range_from_bounds(*unit_b["bounds"]),
                "header_consistent": headers_a == headers_b,
                "headers_a": headers_a, "headers_b": headers_b,
                "rows_a": len(rows_a), "rows_b": len(rows_b),
                "columns": cols,
                "change_count": len(changes),
                "changes": changes,
            })

        overviews_a = [_unit_overview(_pick_unit(units_a[n])) for n in sorted(only_a)]
        overviews_b = [_unit_overview(_pick_unit(units_b[n])) for n in sorted(only_b)]

        total_changes = sum(item["change_count"] for item in matched)
        content = _render_xlsx_report(
            target_a, target_b, matched, overviews_a, overviews_b, table_name, columns)
        content, truncated = _truncate(content)

        metadata = {
            "mode": "xlsx",
            "path_a": str(target_a.resolve()),
            "path_b": str(target_b.resolve()),
            "table_name": table_name,
            "columns": columns,
            "common_count": len(matched),
            "only_in_a": [item["name"] for item in overviews_a],
            "only_in_b": [item["name"] for item in overviews_b],
            "table_diffs": [
                {k: v for k, v in item.items() if k != "changes"} | {
                    "changes": item["changes"][:_MAX_CHANGES_PER_TABLE]}
                for item in matched
            ],
            "overviews_a": overviews_a,
            "overviews_b": overviews_b,
            "total_changes": total_changes,
            "truncated": truncated,
            "warn_threshold": _WARN_THRESHOLD,
        }
        content = _maybe_warn(content, total_changes)
        logger.info("file_diff XLSX完成: 同名表=%d A独有=%d B独有=%d 差异=%d",
                    len(matched), len(overviews_a), len(overviews_b), total_changes)
        return ToolResult(content=content, metadata=metadata)
    finally:
        wb_a.close()
        wb_b.close()


def _collect_named_units(wb) -> dict[str, list[dict]]:
    """收集一个 workbook 内所有"有名数据单元"。

    一个单元 = 一个原生 Excel table（按 table 名），或一张工作表（按 sheet 名）。
    同名（table 名与 sheet 名撞名、或同名表出现在多 sheet）会归到同一 key 下的列表。
    """
    units: dict[str, list[dict]] = {}

    for ws in wb.worksheets:
        # 原生 table
        for table_info in _worksheet_tables(ws):
            name = table_info.get("display_name") or table_info.get("name")
            if not name:
                continue
            try:
                bounds = _parse_a1_range(table_info["ref"])
            except Exception:
                continue
            units.setdefault(name, []).append(
                {"kind": "table", "name": name, "sheet": ws.title, "bounds": bounds, "ws": ws})
        # 工作表本身（用于 sheet 名匹配）
        sheet_bounds = (1, 1, ws.max_column or 1, ws.max_row or 1)
        units.setdefault(ws.title, []).append(
            {"kind": "sheet", "name": ws.title, "sheet": ws.title, "bounds": sheet_bounds, "ws": ws})

    return units


def _pick_unit(unit_list: list[dict]) -> dict:
    """同名多单元时优先取原生 table，否则取第一个。"""
    return sorted(unit_list, key=lambda u: 0 if u["kind"] == "table" else 1)[0]


def _unit_matrix(unit: dict) -> tuple[list[str], list[list[Any]]]:
    """提取单元的 (headers, rows)。headers 为字符串列表，rows 为 JSON 友好值矩阵。"""
    ws = unit["ws"]
    min_col, min_row, max_col, max_row = unit["bounds"]
    if max_row < min_row or max_col < min_col:
        return [], []
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col, values_only=True))
    matrix = [[_json_value(v) for v in row] for row in rows]
    headers = [str(v) if v is not None else "" for v in matrix[0]] if matrix else []
    return headers, matrix


def _resolve_columns(columns: list[int] | None, ncol: int) -> list[int]:
    """解析 columns 参数为 1-based 列索引列表。[a,b] 闭区间，否则显式列。"""
    if not columns:
        return list(range(1, ncol + 1))
    cols = []
    for c in columns:
        try:
            cols.append(int(c))
        except (TypeError, ValueError):
            continue
    if len(cols) == 2 and cols[0] < cols[1]:
        cols = list(range(cols[0], cols[1] + 1))
    return sorted({c for c in cols if 1 <= c <= ncol})


def _diff_matrices(rows_a: list[list[Any]], rows_b: list[list[Any]], cols: list[int]) -> list[dict]:
    """逐行对比两个矩阵（按行号对齐），只比较 cols 指定的列。"""
    changes: list[dict] = []
    max_rows = max(len(rows_a), len(rows_b))
    for ri in range(max_rows):
        a_row = rows_a[ri] if ri < len(rows_a) else None
        b_row = rows_b[ri] if ri < len(rows_b) else None
        for c in cols:
            av = a_row[c - 1] if (a_row and c - 1 < len(a_row)) else None
            bv = b_row[c - 1] if (b_row and c - 1 < len(b_row)) else None
            if av != bv:
                changes.append({"row": ri + 1, "col": c, "a": av, "b": bv})
    return changes


def _unit_overview(unit: dict) -> dict:
    """无同名表时返回的概要：结构 + 样本行。"""
    headers, rows = _unit_matrix(unit)
    data_rows = rows[1:] if rows else []
    return {
        "name": unit["name"],
        "kind": unit["kind"],
        "sheet": unit["sheet"],
        "ref": _range_from_bounds(*unit["bounds"]),
        "data_rows": len(data_rows),
        "cols": len(headers),
        "headers": headers,
        "sample": data_rows[:_SAMPLE_ROWS],
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 报告渲染 + 截断 + 告警
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _render_text_report(
    target_a: Path, target_b: Path, type_a: str, type_b: str,
    len_a: int, len_b: int, stats: dict, hunks: list[dict],
) -> str:
    parts = [
        "== file_diff 报告 ==",
        f"模式: 文本对比 | A: {target_a.name}({type_a}, {len_a} 行) "
        f"| B: {target_b.name}({type_b}, {len_b} 行)",
        "",
        f"【统计】新增 ~{stats['added']} 行 / 删除 ~{stats['removed']} 行 / 变更 hunk {len(hunks)} 个",
        "",
        f"【差异 hunk】(显示前 {_MAX_HUNKS_IN_CONTENT} 个)",
    ]

    if not hunks:
        parts.append("无差异，两文件文本一致。")
        return "\n".join(parts)

    for hunk in hunks[:_MAX_HUNKS_IN_CONTENT]:
        tag = hunk["tag"]
        parts.append(
            f"  @@ A:{hunk['a_start']}-{hunk['a_end']} ↔ B:{hunk['b_start']}-{hunk['b_end']} @@ {tag}")
        for line in hunk["a"]:
            parts.append(f"    - {line}")
        for line in hunk["b"]:
            parts.append(f"    + {line}")

    if len(hunks) > _MAX_HUNKS_IN_CONTENT:
        parts.append(f"  ... 还有 {len(hunks) - _MAX_HUNKS_IN_CONTENT} 个 hunk 未展示")
    return "\n".join(parts)


def _render_xlsx_report(
    target_a: Path, target_b: Path,
    matched: list[dict],
    overviews_a: list[dict], overviews_b: list[dict],
    table_name: str | None, columns: list[int] | None,
) -> str:
    cols_hint = _format_columns_hint(columns)
    parts = [
        "== file_diff 报告 ==",
        f"模式: XLSX 同名表对比 | A: {target_a.name} | B: {target_b.name}"
        + (f" | 限定表: {table_name}" if table_name else "")
        + (f" | 对比列: {cols_hint}" if columns else ""),
        "",
        f"【同名表】共同 {len(matched)} 个；A 独有 {len(overviews_a)} 个；B 独有 {len(overviews_b)} 个",
    ]

    for item in matched:
        parts.append("")
        parts.append(f"【差异 - {item['name']}】(A:{item['sheet_a']} ↔ B:{item['sheet_b']})")
        parts.append(
            f"  表头一致: {'是' if item['header_consistent'] else '否'} | "
            f"行数 A={item['rows_a']} B={item['rows_b']} | 对比列 {item['columns'] or '(空)'}")
        changes = item["changes"]
        if not changes:
            parts.append("  差异: 无（指定列内一致）")
            continue
        shown = changes[:_MAX_CHANGES_PER_TABLE]
        for ch in shown:
            parts.append(f"    - 行{ch['row']} 列{ch['col']}: {_fmt_val(ch['a'])} -> {_fmt_val(ch['b'])}")
        if len(changes) > _MAX_CHANGES_PER_TABLE:
            parts.append(
                f"    ... 共 {len(changes)} 处差异，已展示前 {_MAX_CHANGES_PER_TABLE}。"
                f"收窄: table_name=\"{item['name']}\" columns=[...]")

    for label, overviews in (("A", overviews_a), ("B", overviews_b)):
        for ov in overviews:
            parts.append("")
            parts.append(f"【{label} 独有 - {ov['name']}】(概要)")
            parts.append(f"  sheet={ov['sheet']} ref={ov['ref']} 数据行={ov['data_rows']} 列={ov['cols']}")
            parts.append(f"  表头: {ov['headers']}")
            if ov["sample"]:
                parts.append("  样本:")
                for row in ov["sample"]:
                    parts.append(f"    | {' | '.join(_fmt_val(v) for v in row)}")

    return "\n".join(parts)


def _truncate(content: str) -> tuple[str, bool]:
    """超过 _MAX_OUTPUT_CHARS 时截断并附加提示。"""
    if len(content) <= _MAX_OUTPUT_CHARS:
        return content, False
    kept = content[:_MAX_OUTPUT_CHARS]
    # 尽量截到行尾
    last_nl = kept.rfind("\n")
    if last_nl > _MAX_OUTPUT_CHARS // 2:
        kept = kept[:last_nl]
    kept += (
        "\n\n[已截断] 内容超过上限，仅展示部分。"
        "用 table_name= 指定单个表、或 columns= 限定对比列来获取剩余细节。"
    )
    return kept, True


def _maybe_warn(content: str, total: int) -> str:
    """差异量超阈值时在 content 顶部插告警。"""
    if total <= _WARN_THRESHOLD:
        return content
    warning = (
        f"⚠ 差异量较大：共 {total} 处不一致（告警阈值 {_WARN_THRESHOLD}）。"
        "已截断展示，建议用 table_name/columns 收窄范围。\n\n"
    )
    return warning + content


def _compact_hunks(hunks: list[dict]) -> list[dict]:
    """metadata 里的 hunks 也做截断，避免 metadata 过大。"""
    compacted = []
    for hunk in hunks[:_MAX_HUNKS_IN_CONTENT]:
        compacted.append({
            "tag": hunk["tag"],
            "a_start": hunk["a_start"], "a_end": hunk["a_end"],
            "b_start": hunk["b_start"], "b_end": hunk["b_end"],
            "a": hunk["a"], "b": hunk["b"],
        })
    return compacted


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 小工具
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _normalize_columns(columns: Any) -> list[int] | None:
    if columns is None:
        return None
    if not isinstance(columns, list):
        return None
    out = []
    for c in columns:
        try:
            out.append(int(c))
        except (TypeError, ValueError):
            continue
    return out or None


def _format_columns_hint(columns: list[int] | None) -> str:
    if not columns:
        return ""
    if len(columns) == 2 and columns[0] < columns[1]:
        return f"{columns[0]}~{columns[1]}"
    return ",".join(str(c) for c in columns)


def _fmt_val(value: Any) -> str:
    if value is None:
        return "<空>"
    if isinstance(value, str):
        return value
    return str(value)
