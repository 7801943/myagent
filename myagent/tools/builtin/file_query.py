"""file_query builtin tool.

The tool extracts supported files into located lines, sends line-numbered
chunks to a small subagent model, and returns only answers backed by model
reported evidence lines.
"""
from __future__ import annotations

import asyncio
import csv
import json
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from myagent.tools.api import ToolResult, tool
from myagent.tools.builtin._file_common import (
    _check_path_safety,
    _detect_encoding,
    _detect_file_type,
    _extract_docx_lines,
)
from myagent.utils.config import AgentConfig, load_yaml_config

logger = logging.getLogger(__name__)

_DEFAULT_CONTEXT_WINDOW = 128000
_SUBAGENT_TIMEOUT_SECONDS = 45.0
_SUBAGENT_MAX_TOKENS = 2048
_MAX_CHUNK_INPUT_TOKENS = 32000
_CONTEXT_WINDOW_FRACTION = 0.4
_CHUNK_OVERLAP_LINES = 20
_EXCERPT_CONTEXT_LINES = 2
_SUBAGENT_CONCURRENCY = 2  # chunk 并发查询上限，避免过载 provider


@dataclass(frozen=True)
class LocatedLine:
    global_line: int
    text: str
    source_type: str
    page: int | None = None
    sheet: str | None = None
    local_line: int | None = None


@dataclass(frozen=True)
class LineChunk:
    index: int
    lines: list[LocatedLine]

    @property
    def start_line(self) -> int:
        return self.lines[0].global_line

    @property
    def end_line(self) -> int:
        return self.lines[-1].global_line


@dataclass
class EvidenceRange:
    start_line: int
    end_line: int
    answer: str
    reason: str = ""


class _SimpleMessage:
    """Minimal message object compatible with the existing provider formatters."""

    def __init__(self, role: str, content: str):
        self.role = role
        self.content = content

    def to_openai_dict(self) -> dict[str, str]:
        return {"role": self.role, "content": self.content}

    def to_anthropic_dict(self) -> dict[str, str]:
        return {"role": self.role, "content": self.content}



@tool(
    name="file_query",
    timeout=180,
    description=(
        "使用子代理模型查询大型文件内容。适用于文件过大（比如超过20MB）或者行数过多（比如超过3000行）、直接读取会超过模型最佳上下文区间的场景；"
        "工具会将文件按行自动切块并保留重叠上下文，把每个片段交给子代理判断是否包含答案，"
        "再根据子代理返回的证据行号回填原文引用。支持文本/CSV/DOCX/XLSX/PDF文本层，暂不支持图片、二进制、扫描PDF/OCR。"
        "注意，本工具执行速度较慢，且会消耗大量tokens，在无法直接使用file_read读取时才使用。"
    ),
)
async def file_query(
    path: str,
    query: str,
    mode: str = "answer",
    sheet_name: str | None = None,
    max_evidence: int = 8,
) -> ToolResult:
    """
    从大型文件中查询与 query 相关的信息，并返回答案总结与引用原文。

    Args:
        path: 文件路径。推荐传入绝对路径；如果用户给的是相对路径，调用前请先用当前 workspace root 拼接成绝对路径。本工具不会按 workspace root 自动解析相对路径
        query: 需要在文件中回答的问题或检索目标
        mode: 返回模式，可选 "answer" 或 "evidence"。"answer" 返回答案总结和引用；"evidence" 只返回引用原文
        sheet_name: XLSX 工作表名称。未指定时检索全部工作表
        max_evidence: 最多返回的证据段落数
    """
    logger.info(
        "file_query start: path=%s query_length=%s mode=%s sheet_name=%s max_evidence=%s",
        path,
        len(str(query)) if query is not None else None,
        mode,
        sheet_name,
        max_evidence,
    )
    try:
        result = await _file_query_impl(path, query, mode, sheet_name, max_evidence)
        logger.info(
            "file_query done: is_error=%s evidence_count=%s content_length=%s",
            result.is_error,
            (result.metadata or {}).get("evidence_count"),
            len(result.content or ""),
        )
        return result
    except Exception as exc:
        logger.exception("file_query unexpected error")
        return ToolResult(
            content=f"file_query 执行异常: {type(exc).__name__}: {exc}",
            is_error=True,
            metadata={"type": type(exc).__name__},
        )


async def _file_query_impl(
    path: str | None,
    query: str | None,
    mode: str | None = "answer",
    sheet_name: str | None = None,
    max_evidence: int | None = 8,
) -> ToolResult:
    """
    从文件中检索与 query 相关的信息，并返回答案总结与引用原文。

    Args:
        path: 文件路径。推荐传入绝对路径；如果用户给的是相对路径，调用前请先用当前 workspace root 拼接成绝对路径。本工具不会按 workspace root 自动解析相对路径
        query: 需要在文件中回答的问题或检索目标
        mode: 返回模式，可选 "answer" 或 "evidence"。"answer" 返回答案总结和引用；"evidence" 只返回引用原文
        sheet_name: XLSX 工作表名称。未指定时检索全部工作表
        max_evidence: 最多返回的证据段落数
    """
    path, query, mode, sheet_name, max_evidence = _normalize_inputs(
        path, query, mode, sheet_name, max_evidence
    )
    if not path:
        return ToolResult(content="path 不能为空。", is_error=True)
    if not query:
        return ToolResult(content="query 不能为空。", is_error=True)
    if mode not in ("answer", "evidence"):
        return ToolResult(content=f"不支持的 mode: {mode}。可选: answer, evidence", is_error=True)

    error = _check_path_safety(path)
    if error:
        return ToolResult(content=error, is_error=True)

    target = Path(path)
    if not target.exists():
        return ToolResult(content=f"文件不存在: {path}", is_error=True)
    if not target.is_file():
        return ToolResult(content=f"不是文件: {path}", is_error=True)

    file_type = _detect_file_type(target)
    if file_type in ("image", "binary"):
        return ToolResult(
            content=f"file_query 不支持 {file_type} 文件；首版仅支持文本/CSV/DOCX/XLSX/PDF文本层。",
            is_error=True,
        )

    try:
        lines = await _extract_located_lines(target, file_type, sheet_name)
        logger.info("file_query extracted: file_type=%s line_count=%s", file_type, len(lines))
    except Exception as exc:
        logger.exception("file_query extract failed")
        return ToolResult(content=f"解析文件失败: {type(exc).__name__}: {exc}", is_error=True)

    if not lines:
        return ToolResult(
            content="文件未提取到可检索文本。",
            metadata={"path": str(target.resolve()), "format": file_type, "query": query, "mode": mode},
        )

    context_window = _get_primary_context_window()
    chunk_token_limit = int(min(_MAX_CHUNK_INPUT_TOKENS, max(1000, context_window * _CONTEXT_WINDOW_FRACTION)))
    chunks = _chunk_lines(lines, chunk_token_limit, _CHUNK_OVERLAP_LINES)
    logger.info(
        "file_query chunked: chunk_count=%s first_chunk=%s last_chunk=%s",
        len(chunks),
        (chunks[0].start_line, chunks[0].end_line) if chunks else None,
        (chunks[-1].start_line, chunks[-1].end_line) if chunks else None,
    )
    if not chunks:
        return ToolResult(content="文件未生成可发送给子代理的文本 chunk。", is_error=True)

    evidence_ranges: list[EvidenceRange] = []
    chunk_failures: list[dict[str, Any]] = []
    chunks_with_answer = 0

    # 在入口处构建一次 router，供所有 chunk 复用，避免每个 chunk 重新读 config
    router = _build_subagent_router()
    semaphore = asyncio.Semaphore(_get_subagent_concurrency())

    async def _process_chunk(chunk: LineChunk) -> None:
        nonlocal chunks_with_answer
        async with semaphore:
            try:
                parsed = await _query_chunk_with_retry(
                    query=query, chunk=chunk, router=router
                )
            except Exception as exc:
                logger.exception("file_query subagent chunk failed: chunk=%s", chunk.index)
                failure_kind = "invalid_json" if isinstance(exc, ValueError) else "subagent_error"
                chunk_failures.append({
                    "chunk": chunk.index,
                    "kind": failure_kind,
                    "error": f"{type(exc).__name__}: {exc}",
                })
                return

        valid = _evidence_from_chunk_result(parsed, chunk)
        if not valid:
            return
        chunks_with_answer += 1
        evidence_ranges.extend(valid)

    await asyncio.gather(*(_process_chunk(chunk) for chunk in chunks))

    if not evidence_ranges and chunk_failures and len(chunk_failures) == len(chunks):
        failure_summary = _summarize_chunk_failures(chunk_failures)
        return ToolResult(
            content=(
                "file_query 子代理不可用或未能完成检索，所有 chunk 均调用失败或返回无效结果。"
                f"首个错误: {chunk_failures[0]['error']}"
            ),
            is_error=True,
            metadata={
                "path": str(target.resolve()),
                "format": file_type,
                "query": query,
                "mode": mode,
                "chunk_count": len(chunks),
                "chunk_failures": chunk_failures,
                "failure_summary": failure_summary,
            },
        )

    merged = _merge_evidence(evidence_ranges)[:max_evidence]
    evidence_payload = _build_evidence_payload(merged, lines, _EXCERPT_CONTEXT_LINES)
    answer_text = _build_answer_text(mode, merged, evidence_payload)
    content = _format_tool_content(answer_text, evidence_payload)

    metadata = {
        "path": str(target.resolve()),
        "format": file_type,
        "query": query,
        "mode": mode,
        "chunk_count": len(chunks),
        "chunks_with_answer": chunks_with_answer,
        "evidence_count": len(evidence_payload),
        "evidence": evidence_payload,
    }
    if chunk_failures:
        metadata["chunk_failures"] = chunk_failures

    return ToolResult(content=content, metadata=metadata)


def _normalize_inputs(
    path: Any,
    query: Any,
    mode: Any,
    sheet_name: Any,
    max_evidence: Any,
) -> tuple[str, str, str, str | None, int]:
    path_text = "" if path is None else str(path).strip()
    query_text = "" if query is None else str(query).strip()
    mode_text = "answer" if mode is None else str(mode).strip().lower() or "answer"
    sheet_text = None if sheet_name is None else str(sheet_name).strip() or None
    evidence_count = _coerce_int(max_evidence, 8, minimum=1, maximum=50)
    return path_text, query_text, mode_text, sheet_text, evidence_count


def _coerce_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        value = int(value)
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(maximum, value))


async def _extract_located_lines(path: Path, file_type: str, sheet_name: str | None) -> list[LocatedLine]:
    if file_type == "text":
        return _extract_text_lines(path)
    if file_type == "csv":
        return _extract_csv_lines(path)
    if file_type == "docx":
        return _extract_docx_located_lines(path)
    if file_type == "xlsx":
        return _extract_xlsx_lines(path, sheet_name)
    if file_type == "pdf":
        return _extract_pdf_text_lines(path)
    raise ValueError(f"不支持的文件类型: {file_type}")


def _extract_text_lines(path: Path) -> list[LocatedLine]:
    enc = _detect_encoding(path)
    if enc == "binary":
        raise ValueError("文件似乎是二进制文件")
    with open(path, "r", encoding=enc, errors="replace") as f:
        raw_lines = f.readlines()
    return [
        LocatedLine(i, line.rstrip("\n\r"), source_type="text", local_line=i)
        for i, line in enumerate(raw_lines, 1)
    ]


def _extract_csv_lines(path: Path) -> list[LocatedLine]:
    enc = _detect_encoding(path)
    if enc == "binary":
        enc = "utf-8"
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, "r", encoding=enc, errors="replace", newline="") as f:
        rows = list(csv.reader(f, delimiter=delimiter))
    return [
        LocatedLine(i, " | ".join(str(cell) for cell in row), source_type="csv", local_line=i)
        for i, row in enumerate(rows, 1)
    ]


def _extract_docx_located_lines(path: Path) -> list[LocatedLine]:
    if path.suffix.lower() == ".doc":
        raise ValueError("不支持旧版 .doc 格式，请先转换为 .docx")
    from docx import Document

    doc = Document(str(path))
    raw_lines = _extract_docx_lines(doc)
    return [
        LocatedLine(i, line.rstrip("\n\r"), source_type="docx", local_line=i)
        for i, line in enumerate(raw_lines, 1)
    ]


def _extract_xlsx_lines(path: Path, sheet_name: str | None) -> list[LocatedLine]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet_names = [sheet_name] if sheet_name else wb.sheetnames
        missing = [name for name in sheet_names if name not in wb.sheetnames]
        if missing:
            raise ValueError(f"工作表不存在: {missing[0]}。可用工作表: {', '.join(wb.sheetnames)}")

        located: list[LocatedLine] = []
        global_line = 1
        for name in sheet_names:
            ws = wb[name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = ["" if value is None else str(value) for value in row]
                text = " | ".join(cells).rstrip()
                # 跳过空行，避免浪费子代理 token（保留 local_line 行号映射）
                if not text:
                    continue
                located.append(
                    LocatedLine(
                        global_line=global_line,
                        text=text,
                        source_type="xlsx",
                        sheet=name,
                        local_line=row_idx,
                    )
                )
                global_line += 1
        return located
    finally:
        wb.close()


def _extract_pdf_text_lines(path: Path) -> list[LocatedLine]:
    for logger_name in (
        "pdfminer",
        "pdfminer.psparser",
        "pdfminer.pdfinterp",
        "pdfminer.pdfpage",
        "pdfminer.converter",
        "pdfminer.layout",
        "pdfminer.utils",
    ):
        logging.getLogger(logger_name).setLevel(logging.WARNING)

    import pdfplumber

    located: list[LocatedLine] = []
    global_line = 1
    with pdfplumber.open(str(path)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            if not text.strip():
                continue
            for page_line, line in enumerate(text.split("\n"), 1):
                located.append(
                    LocatedLine(
                        global_line=global_line,
                        text=line.rstrip("\n\r"),
                        source_type="pdf",
                        page=page_num,
                        local_line=page_line,
                    )
                )
                global_line += 1
    return located


def _get_primary_context_window(config_path: str = "config.yaml") -> int:
    try:
        raw = load_yaml_config(config_path)
        app_config = raw.get("agent", raw) if raw else {}
        cfg = AgentConfig(**app_config)
        if cfg.providers:
            primary = sorted(cfg.providers, key=lambda provider: provider.priority)[0]
            return primary.context_window_size
    except Exception as exc:
        logger.exception("file_query failed to read provider context_window_size, using default")
    return _DEFAULT_CONTEXT_WINDOW


def _get_subagent_concurrency() -> int:
    try:
        value = int(os.environ.get("MYAGENT_FILE_QUERY_CONCURRENCY", _SUBAGENT_CONCURRENCY))
    except (TypeError, ValueError):
        value = _SUBAGENT_CONCURRENCY
    return max(1, min(8, value))


def _summarize_chunk_failures(chunk_failures: list[dict[str, Any]]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for item in chunk_failures:
        kind = str(item.get("kind") or "unknown")
        summary[kind] = summary.get(kind, 0) + 1
    return summary


def _chunk_lines(lines: list[LocatedLine], token_limit: int, overlap_lines: int) -> list[LineChunk]:
    chunks: list[LineChunk] = []
    current: list[LocatedLine] = []
    current_tokens = 0

    for line in lines:
        line_tokens = max(1, int((len(line.text) + 16) / 3))
        if current and current_tokens + line_tokens > token_limit:
            chunks.append(LineChunk(index=len(chunks) + 1, lines=list(current)))
            overlap = current[-overlap_lines:] if overlap_lines else []
            current = list(overlap)
            current_tokens = sum(max(1, int((len(item.text) + 16) / 3)) for item in current)
        current.append(line)
        current_tokens += line_tokens

    if current:
        chunks.append(LineChunk(index=len(chunks) + 1, lines=list(current)))
    return chunks


async def _query_chunk_with_retry(
    query: str,
    chunk: LineChunk,
    router: Any = None,
) -> dict[str, Any]:
    last_error: Exception | None = None
    invalid_response = ""
    for attempt in range(2):
        prompt = _build_chunk_prompt(query, chunk, invalid_response=invalid_response)
        messages = [
            _SimpleMessage(
                "system",
                (
                    "你是文件内容检索agent。你只能依据用户提供的当前文件片段回答。"
                    "如果片段没有答案，必须返回 has_answer=false 和空 evidence。"
                    "只输出 JSON，不要输出 Markdown。"
                ),
            ),
            _SimpleMessage("user", prompt),
        ]
        # router 复用：如果传入了预构建的 router，则不再每次重新读 config
        raw = await _call_subagent_model(messages, router=router)
        try:
            parsed = _parse_json_object(raw)
            return parsed
        except ValueError as exc:
            last_error = exc
            invalid_response = raw[:1000]
            logger.exception("file_query subagent JSON parse failed: attempt=%s chunk=%s", attempt + 1, chunk.index)
    raise ValueError(f"子代理未返回合法 JSON: {last_error}")


def _build_chunk_prompt(query: str, chunk: LineChunk, invalid_response: str = "") -> str:
    lines_text = "\n".join(f"{line.global_line} | {line.text}" for line in chunk.lines)
    retry_hint = ""
    if invalid_response:
        retry_hint = (
            "\n上一次返回不是合法 JSON，请修正。上一次返回片段如下：\n"
            f"{invalid_response}\n"
        )
    return (
        f"问题：{query}\n\n"
        f"当前文件片段行号范围：{chunk.start_line}-{chunk.end_line}\n"
        "请判断这个片段是否包含能回答问题的依据。证据行号必须使用左侧的全局行号。\n"
        "返回 JSON schema：\n"
        '{"has_answer": true, "answer": "基于本段证据的简短答案", '
        '"evidence": [{"start_line": 10, "end_line": 14, "reason": "这里包含答案依据"}]}\n'
        "如果没有证据，返回：\n"
        '{"has_answer": false, "answer": "", "evidence": []}\n'
        f"{retry_hint}\n"
        "文件片段：\n"
        f"{lines_text}"
    )


def _build_subagent_router(config_path: str = "config.yaml") -> Any:
    """构建子代理 ProviderRouter。在入口调用一次，供所有 chunk 复用。

    Raises:
        RuntimeError: 未配置任何 Provider 或未找到支持的 Provider。
    """
    raw = load_yaml_config(config_path)
    app_config = raw.get("agent", raw) if raw else {}
    cfg = AgentConfig(**app_config)
    if not cfg.providers:
        raise RuntimeError("未配置任何 Provider，file_query 子代理不可用")

    providers = []
    for provider_cfg in sorted(cfg.providers, key=lambda item: item.priority):
        provider_type = provider_cfg.type.lower()
        if provider_type == "openai":
            from myagent.providers.openai_provider import OpenAIProvider

            provider = OpenAIProvider(
                name=provider_cfg.name,
                model=provider_cfg.model,
                api_key=provider_cfg.api_key or "sk-dummy",
                api_base=provider_cfg.api_base,
            )
        elif provider_type == "anthropic":
            from myagent.providers.anthropic_provider import AnthropicProvider

            provider = AnthropicProvider(
                name=provider_cfg.name,
                model=provider_cfg.model,
                api_key=provider_cfg.api_key or "sk-dummy",
                api_base=provider_cfg.api_base,
            )
        else:
            continue
        provider._priority = provider_cfg.priority
        providers.append(provider)

    if not providers:
        raise RuntimeError("未找到支持的 Provider，file_query 子代理不可用")

    from myagent.providers.router import ProviderRouter

    router = ProviderRouter(
        providers,
        failure_threshold=cfg.failover.circuit_breaker_failure_threshold,
        recovery_seconds=cfg.failover.circuit_breaker_recovery_seconds,
    )
    return router


async def _call_subagent_model(
    messages: list[_SimpleMessage],
    config_path: str = "config.yaml",
    max_tokens: int = _SUBAGENT_MAX_TOKENS,
    timeout: float = _SUBAGENT_TIMEOUT_SECONDS,
    router: Any = None,
) -> str:
    # router 复用：如果传入了预构建的 router，则跳过 config 读取和 router 构建
    if router is None:
        router = _build_subagent_router(config_path)

    async def collect() -> str:
        parts: list[str] = []
        async for event in router.stream(messages, tools=None, max_tokens=max_tokens):
            if event.type == "text_delta" and event.text:
                parts.append(event.text)
        return "".join(parts)

    try:
        result = await asyncio.wait_for(collect(), timeout=timeout)
        return result
    except Exception:
        logger.exception("file_query subagent collect failed")
        raise


def _parse_json_object(raw: str) -> dict[str, Any]:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start < 0 or end <= start:
            raise ValueError("响应中未找到 JSON 对象")
        parsed = json.loads(text[start : end + 1])
    if not isinstance(parsed, dict):
        raise ValueError("JSON 顶层必须是对象")
    return parsed


def _evidence_from_chunk_result(parsed: dict[str, Any], chunk: LineChunk) -> list[EvidenceRange]:
    if not parsed.get("has_answer"):
        return []
    evidence = parsed.get("evidence")
    if not isinstance(evidence, list) or not evidence:
        return []

    answer = str(parsed.get("answer") or "").strip()
    chunk_lines = {line.global_line for line in chunk.lines}
    ranges: list[EvidenceRange] = []

    for item in evidence:
        if not isinstance(item, dict):
            continue
        try:
            start_line = int(item.get("start_line"))
            end_line = int(item.get("end_line"))
        except (TypeError, ValueError):
            continue
        if start_line > end_line:
            continue
        if start_line not in chunk_lines or end_line not in chunk_lines:
            continue
        reason = str(item.get("reason") or "").strip()
        ranges.append(EvidenceRange(start_line=start_line, end_line=end_line, answer=answer, reason=reason))
    return ranges


def _merge_evidence(evidence: list[EvidenceRange]) -> list[EvidenceRange]:
    if not evidence:
        return []
    ordered = sorted(evidence, key=lambda item: (item.start_line, item.end_line))
    merged: list[EvidenceRange] = []

    for item in ordered:
        if not merged:
            merged.append(item)
            continue
        prev = merged[-1]
        if item.start_line <= prev.end_line + 1:
            prev.end_line = max(prev.end_line, item.end_line)
            if item.answer and item.answer not in prev.answer:
                prev.answer = _join_unique_text(prev.answer, item.answer)
            if item.reason and item.reason not in prev.reason:
                prev.reason = _join_unique_text(prev.reason, item.reason)
        else:
            merged.append(item)
    return merged


def _join_unique_text(left: str, right: str) -> str:
    if not left:
        return right
    if not right:
        return left
    return f"{left}；{right}"


def _build_evidence_payload(
    ranges: list[EvidenceRange],
    lines: list[LocatedLine],
    context_lines: int,
) -> list[dict[str, Any]]:
    by_global = {line.global_line: line for line in lines}
    max_line = max(by_global) if by_global else 0
    payload: list[dict[str, Any]] = []

    for index, item in enumerate(ranges, 1):
        start = max(1, item.start_line - context_lines)
        end = min(max_line, item.end_line + context_lines)
        excerpt_lines = [by_global[i] for i in range(start, end + 1) if i in by_global]
        evidence_lines = [by_global[i] for i in range(item.start_line, item.end_line + 1) if i in by_global]
        if not excerpt_lines or not evidence_lines:
            continue

        locator = _format_locator(evidence_lines, item.start_line, item.end_line)
        text = "\n".join(f"{line.global_line} | {line.text}" for line in excerpt_lines)
        first = evidence_lines[0]
        payload.append(
            {
                "source_id": f"S{index}",
                "locator": locator,
                "start_line": item.start_line,
                "end_line": item.end_line,
                "page": first.page,
                "sheet": first.sheet,
                "reason": item.reason,
                "answer": item.answer,
                "text": text,
            }
        )
    return payload


def _format_locator(evidence_lines: list[LocatedLine], start_line: int, end_line: int) -> str:
    pages = {line.page for line in evidence_lines if line.page is not None}
    sheets = {line.sheet for line in evidence_lines if line.sheet is not None}
    local_lines = [line.local_line for line in evidence_lines if line.local_line is not None]

    if len(pages) == 1 and local_lines:
        page = next(iter(pages))
        return f"page {page}, L{min(local_lines)}-L{max(local_lines)}"
    if len(sheets) == 1 and local_lines:
        sheet = next(iter(sheets))
        return f"{sheet}: L{min(local_lines)}-L{max(local_lines)}"
    return f"L{start_line}-L{end_line}"


def _build_answer_text(mode: str, evidence_ranges: list[EvidenceRange], evidence_payload: list[dict[str, Any]]) -> str:
    if not evidence_payload:
        return "未在文件中找到足够依据。"
    if mode == "evidence":
        return "（mode=evidence，未生成答案总结。）"

    answer_rows: list[str] = []
    for item in evidence_payload:
        answer = str(item.get("answer") or "").strip()
        if not answer:
            continue
        source_id = item["source_id"]
        row = f"- {answer} [{source_id}]"
        if row not in answer_rows:
            answer_rows.append(row)

    if not answer_rows and evidence_ranges:
        return "已找到相关证据，但子代理未返回可保留的答案总结。"
    return "\n".join(answer_rows) if answer_rows else "未在文件中找到足够依据。"


def _format_tool_content(answer_text: str, evidence_payload: list[dict[str, Any]]) -> str:
    parts = ["答案总结", answer_text, "", "引用原文"]
    if not evidence_payload:
        parts.append("无。")
        return "\n".join(parts)

    for item in evidence_payload:
        parts.append(f"[{item['source_id']}] {item['locator']}")
        if item.get("reason"):
            parts.append(f"原因: {item['reason']}")
        parts.append("```")
        parts.append(str(item["text"]))
        parts.append("```")
    return "\n".join(parts)
