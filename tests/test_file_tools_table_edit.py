import asyncio

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from myagent.tools.builtin.file_edit import file_edit_table
from myagent.tools.builtin.file_read import file_read


def run_tool(coro):
    return asyncio.run(coro)


def save_basic_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Code"
    ws["B1"] = "Amount"
    wb.save(path)


def save_table_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["OrderId", "Status", "Amount"])
    ws.append(["A-1", "Open", 10])
    table = Table(displayName="SalesTable", ref="A1:C2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(path)


def test_set_range_defaults_to_dry_run_and_preserves_text_formula(tmp_path):
    path = tmp_path / "basic.xlsx"
    save_basic_workbook(path)

    preview = run_tool(file_edit_table(
        str(path),
        operation="set_range",
        sheet_name="Data",
        payload={"range": "A2:B2", "values": [["00123", "=10+5"]]},
        include_changes=True,
    ))

    assert not preview.is_error
    assert preview.metadata["dry_run"] is True
    assert preview.metadata["data"]["changes"][0]["new_value"] == "00123"

    wb = load_workbook(path, data_only=False)
    assert wb["Data"]["A2"].value is None
    wb.close()

    result = run_tool(file_edit_table(
        str(path),
        operation="set_range",
        sheet_name="Data",
        payload={"range": "A2:B2", "values": [["00123", "=10+5"]]},
        dry_run=False,
    ))

    assert not result.is_error, result.content
    assert "值变更 diff" in result.content
    assert "A2" in result.metadata["data"]["diff_preview"]
    assert result.metadata["data"]["persistence_verification"]["persisted"] is True
    assert result.metadata["data"]["change_count"] == 2
    wb = load_workbook(path, data_only=False)
    assert wb["Data"]["A2"].value == "00123"
    assert wb["Data"]["B2"].value == "=10+5"
    wb.close()



def test_format_range_reports_non_value_result_and_verifies_persistence(tmp_path):
    path = tmp_path / "format.xlsx"
    save_basic_workbook(path)

    result = run_tool(file_edit_table(
        str(path),
        operation="format_range",
        sheet_name="Data",
        payload={
            "range": "A1:B1",
            "fill": "#FFF2CC",
            "font": {"bold": True},
            "comment": "reviewed",
        },
        dry_run=False,
    ))

    assert not result.is_error, result.content
    assert "值变更 diff: 无单元格值变化" in result.content
    assert "格式/批注结果" in result.content
    assert result.metadata["data"]["format_actions"][0]["fill"] == "FFF2CC"
    assert result.metadata["data"]["format_actions"][0]["comment"] == "reviewed"
    assert result.metadata["data"]["persistence_verification"]["persisted"] is True
    assert result.metadata["data"]["persistence_verification"]["format_cells_checked"] == 2

    wb = load_workbook(path, data_only=False)
    ws = wb["Data"]
    assert ws["A1"].fill.fgColor.rgb.endswith("FFF2CC")
    assert ws["A1"].comment.text == "reviewed"
    assert ws["B1"].font.bold is True
    wb.close()


def test_update_cells_and_clear_range(tmp_path):
    path = tmp_path / "cells.xlsx"
    save_basic_workbook(path)

    result = run_tool(file_edit_table(
        str(path),
        operation="update_cells",
        sheet_name="Data",
        payload={
            "cells": [
                {"cell": "A2", "value": "00042", "kind": "text"},
                {"cell": "B2", "value": 42, "kind": "number"},
            ]
        },
        dry_run=False,
    ))
    assert not result.is_error, result.content

    result = run_tool(file_edit_table(
        str(path),
        operation="clear_range",
        sheet_name="Data",
        payload={"range": "B2:B2"},
        dry_run=False,
    ))
    assert not result.is_error, result.content

    wb = load_workbook(path, data_only=False)
    assert wb["Data"]["A2"].value == "00042"
    assert wb["Data"]["B2"].value is None
    wb.close()


def test_append_rows_and_update_rows_by_key_for_native_table(tmp_path):
    path = tmp_path / "sales.xlsx"
    save_table_workbook(path)

    appended = run_tool(file_edit_table(
        str(path),
        operation="append_rows",
        payload={
            "table_name": "SalesTable",
            "rows": [{"OrderId": "A-2", "Status": "Open", "Amount": 25}],
        },
        dry_run=False,
        allow_structure_change=True,
    ))
    assert not appended.is_error, appended.content

    wb = load_workbook(path, data_only=False)
    ws = wb["Sales"]
    assert ws.tables["SalesTable"].ref == "A1:C3"
    assert ws["A3"].value == "A-2"
    wb.close()

    updated = run_tool(file_edit_table(
        str(path),
        operation="update_rows_by_key",
        payload={
            "table_name": "SalesTable",
            "key_column": "OrderId",
            "updates": [{"key": "A-1", "values": {"Status": "Paid"}}],
        },
        dry_run=False,
    ))
    assert not updated.is_error, updated.content

    wb = load_workbook(path, data_only=False)
    assert wb["Sales"]["B2"].value == "Paid"
    wb.close()


def test_delete_rows_refuses_formula_sheet_even_with_structure_flag(tmp_path):
    path = tmp_path / "formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(path)

    result = run_tool(file_edit_table(
        str(path),
        operation="delete_rows",
        sheet_name="Data",
        payload={"rows": [1]},
        dry_run=False,
        allow_structure_change=True,
    ))

    assert result.is_error
    assert "公式" in result.content

    wb = load_workbook(path, data_only=False)
    assert wb["Data"]["A1"].value == 1
    assert wb["Data"]["A2"].value == "=A1+1"
    wb.close()


def test_file_read_xlsx_returns_structured_metadata(tmp_path):
    path = tmp_path / "read.xlsx"
    save_basic_workbook(path)
    run_tool(file_edit_table(
        str(path),
        operation="set_range",
        sheet_name="Data",
        payload={"range": "A2:B2", "values": [["A-1", "=1+1"]]},
        dry_run=False,
    ))

    result = run_tool(file_read(
        str(path),
        sheet_name="Data",
        xlsx_range="A1:B2",
        render_mode="both",
        row_mode="arrays",
    ))

    assert not result.is_error, result.content
    assert result.metadata["range"] == "A1:B2"
    assert result.metadata["row_mode"] == "arrays"
    assert result.metadata["structure_token"]
    assert result.metadata["content_token"]
    assert result.metadata["rows"][1][0] == "A-1"
    assert result.metadata["rows"][1][1]["formula"] == "=1+1"
